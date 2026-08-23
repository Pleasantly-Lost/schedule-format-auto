from ctypes import wstring_at
from tkinter.messagebox import showerror

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.styles import Font
from datetime import time, datetime, timedelta
from pathlib import Path
import pyexcel_io.writers
import pyexcel_io.readers
import pyexcel as p
import os
import re
from xlrd.formula import sheetrange


# Global constants
NORMAL_ROUTES = {
    "ER-01", "SR-02", "DR-03A", "DR-03B", "DR-05", "DR-06", "DR-07", "SR-08",
    "ER-09", "DR-11", "DR-13", "XER-15", "ER-16", "DR-17", "DR-18"
}
QUEER_ROUTES = {"ER-10", "ER-12", "DR-04B", "DR-014", "DR-014A", }
ALIASED_PREFIXES = {"EXP": "ER"}


def copy_column_as_text_between_workbooks(
    source_wb_path,
    source_sheet_index,
    source_cell_ref,
    target_wb_path,
    target_sheet_index,
    target_cell_ref,
    convert_time_to_hhmm=False
):
    # Load both workbooks
    source_wb = load_workbook(source_wb_path, data_only=True)  # data_only=True avoids formulas
    target_wb = load_workbook(target_wb_path)


    # Get worksheets
    ws_source = source_wb.worksheets[source_sheet_index]
    ws_target = target_wb.worksheets[target_sheet_index]

    # Parse source and target cell references
    source_col_letter, source_start_row = coordinate_from_string(source_cell_ref)
    target_col_letter, target_start_row = coordinate_from_string(target_cell_ref)

    # Step 1: Find the last non-empty cell in the source column
    current_row = source_start_row
    while True:
        if ws_source[f"{source_col_letter}{current_row}"].value is None:
            break
        current_row += 1
    end_row = current_row - 1

    # Step 2: Copy values to target, formatting as Text
    for i, row in enumerate(ws_source[f"{source_col_letter}{source_start_row}:{source_col_letter}{end_row}"]):
        source_cell = row[0]
        value = source_cell.value
        target_row = target_start_row + i
        target_cell = ws_target[f"{target_col_letter}{target_row}"]

        if convert_time_to_hhmm and isinstance(value, time):
            # Format time as HH:MM and save as string
            time_str = value.strftime('%H:%M')
            target_cell.value = time_str
            target_cell.number_format = '@'
        else:
            # Save other values as string with text format
            target_cell.value = str(value) if value is not None else ""
            target_cell.number_format = '@'

    # Save the target workbook only
    target_wb.save(target_wb_path)

def autofill_next_row(file_path:str):
    # 1. Load the workbook
    wb = load_workbook(file_path)
    ws = wb.worksheets[0]

    last_row_idx = ws.max_row
    if last_row_idx < 2:
        raise ValueError("The worksheet needs at least a header and one data row.")

    # 3. Extract values from the very last populated row
    last_shift = ws.cell(row=last_row_idx, column=1).value  # Col A: Shift Number
    shift_type = ws.cell(row=last_row_idx, column=2).value  # Col B: Shift Type
    scheme = ws.cell(row=last_row_idx, column=3).value  # Col C: Scheme
    last_departure = ws.cell(row=last_row_idx, column=4).value  # Col D: Departure Time
    min_duration = ws.cell(row=last_row_idx, column=5).value  # Col E: Min Trip
    max_duration = ws.cell(row=last_row_idx, column=6).value  # Col F: Max Trip

    # 4. Calculate your dynamic autofill increments
    next_shift = int(last_shift) + 1

    # Safely convert time string to increment exactly 1 minute
    time_obj = datetime.strptime(str(last_departure).strip(), "%H:%M")
    next_departure = (time_obj + timedelta(minutes=1)).strftime("%H:%M")

    # 5. Compile the new row array
    new_row_data = [
        next_shift,  # Column A
        shift_type,  # Column B
        scheme,  # Column C
        next_departure,  # Column D
        min_duration,  # Column E
        max_duration  # Column F
    ]

    # 6. Append cleanly to the bottom of the worksheet and save
    ws.append(new_row_data)
    wb.save(file_path)
    wb.close()

    print(f"Row {last_row_idx + 1} added to sheet '{ws.title}'. Shift: {next_shift}, Time: {next_departure}")


def get_travel_time(source_file, sheet_index, target):

    wb_source = openpyxl.load_workbook(source_file)

    if sheet_index >=  len(wb_source.sheetnames) or sheet_index < 0:
        print(f"Error: Sheet  index'{sheet_index}' is out of range.")
        return[]

    target_sheet_name = wb_source.sheetnames[sheet_index]
    ws_source = wb_source[target_sheet_name]

    search_term = str(target).lower()
    found_instances = []

    for row in ws_source.iter_rows():
        for cell in row:
            if cell.value and search_term in str(cell.value).lower():

                offset_col = 1
                target_cell = cell.offset(row=0, column=offset_col)

                while target_cell.value is None:
                    offset_col += 1
                    if offset_col > 10:
                        break
                    target_cell = cell.offset(row=0, column=offset_col)

                if target_cell.value is not None:

                    found_instances.append(
                        {
                            "coordinate": target_cell.coordinate,
                            "column_index": target_cell.column
                        }
                    )

    found_instances = sorted(found_instances, key=lambda x: x["column_index"])
    return found_instances

def convert_to_total_minutes(time_obj) -> str:
    if time_obj is None:
        return "0"

    total_minutes = (time_obj.hour * 60) + time_obj.minute

    return str(total_minutes)

def get_header_coordinates(source_file, sheet_index, target_header):

    wb_source = openpyxl.load_workbook(source_file, read_only=True)

    if sheet_index >=  len(wb_source.sheetnames) or sheet_index < 0:
        print(f"Error: Sheet  index'{sheet_index}' is out of range.")
        return[]

    target_sheet_name = wb_source.sheetnames[sheet_index]
    ws_source = wb_source[target_sheet_name]

    print(f"Scanning Sheet Index {sheet_index} (Named: '{target_sheet_name}')...")

    search_term = str(target_header).lower()
    found_instances = []

    for row in ws_source.iter_rows():
        for cell in row:
            if cell.value and search_term in str(cell.value).lower():
                found_instances.append(
                    {
                        "coordinate": cell.coordinate,
                        "column_letter": cell.column_letter,
                        "data row": cell.column_letter + str(cell.row + 1),
                        "column_index": cell.column,
                    }
                )
    return found_instances

def trim_empty_rows_and_columns(filename):
    wb = load_workbook(filename)

    for ws in wb.worksheets:
        # Find last row and column with data
        max_row = 0
        max_col = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    max_row = max(max_row, cell.row)
                    max_col = max(max_col, cell.column)

        # Delete rows below the last data row
        if max_row < ws.max_row:
            ws.delete_rows(max_row + 1, ws.max_row - max_row)

        # Delete columns after the last data column
        if max_col < ws.max_column:
            ws.delete_cols(max_col + 1, ws.max_column - max_col)

    wb.save(filename)

def write_header(sheet, header):
    for col_num, header_value in enumerate(header, start=1):
        cell = sheet.cell(row=1, column=col_num, value=header_value)
        cell.font = Font(bold=True)

def convert_xlsx_to_xls(xlsx_file, xls_file):
    # Read the .xlsx file
    sheet = p.get_book(file_name=xlsx_file)

    # Save as .xls
    sheet.save_as(xls_file)
    
def scheme_and_trip_duration (
        source_wb_path,
        out_wb_path,
        fwd_trip_duration,
        bwd_trip_duration
):
    wb_source = load_workbook(source_wb_path)
    wb_out = load_workbook(out_wb_path)

    shift_change_time = datetime.strptime("14:00", "%H:%M").time()

    for idx, sheet in enumerate(wb_out.worksheets):

        max_row = 0
        for row in range(1, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value or sheet.cell(row=row, column=4).value:
                max_row = row

        for row in range(2, max_row + 1):
            time_str = sheet.cell(row=row, column=4).value

            shift = ""

            if isinstance(time_str, str):
                try:
                    time_obj = datetime.strptime(time_str.strip(), "%H:%M").time()
                    shift = "Morning shift" if time_obj < shift_change_time else "Night shift"
                except ValueError:
                    shift = "ERROR!"

            sheet.cell(row=row, column=2).value = shift
            sheet.cell(row=row, column=5).value = fwd_trip_duration if idx == 0 else bwd_trip_duration
            sheet.cell(row=row, column=6).value = fwd_trip_duration if idx == 0 else bwd_trip_duration
            sheet.cell(row=row, column=3).value = "Forward" if idx == 0 else "Backward"

    wb_out.save(out_wb_path)

def get_route_name(input_path: str) -> str:
    filename_stem = Path(input_path).stem.upper()

    match = re.search(r'(EXP|SR|DR|ER|XER)[-\s]?(\d+)\s*([AB]?)', filename_stem)
    if not match:
        raise ValueError(f"No valid route found in filename '{Path(input_path).name}' .")

    prefix, num_str, suffix = match.group(1), match.group(2), match.group(3)

    prefix = ALIASED_PREFIXES.get(prefix, prefix)

    if prefix == "DR" and num_str in ("14", "014"):
        normalized_num = num_str.zfill(3)
    else:
        normalized_num = num_str.zfill(2)

    final_output_route = f"{prefix}-{normalized_num}{suffix}"

    if final_output_route not in NORMAL_ROUTES and final_output_route not in QUEER_ROUTES:
        raise ValueError(f"Extracted route '{final_output_route}' is unrecognized.")

    return final_output_route

def run_excel_stuff(
        input_path_from_user
):

    # Preparing input worksheet -
    path = input_path_from_user # here is the input file path - hook up to GUI
    just_the_name = os.path.basename(path)

    identified_route = get_route_name(path)
    print(f"Route detected as: {identified_route}")

    # Preparing input woksheets
    wb_in = openpyxl.load_workbook(path, data_only=True) # why is only this one got data_only=True ?
    ws_stats = wb_in.worksheets[0]
    ws_details = wb_in.worksheets[1]

    # Preparing output worksheet
    wb_out = Workbook()

    wb_out_forward_sheet = identified_route + "(Forward)"
    wb_out_backward_sheet = identified_route + "(Backward)"

    wb_out_forward = wb_out.create_sheet(wb_out_forward_sheet)
    wb_out_backward = wb_out.create_sheet(wb_out_backward_sheet)

    if 'Sheet' in wb_out.sheetnames:
        del wb_out['Sheet']

    # Write Headers (IMP for schedule)
    schedule_header = ["Shift Number","Shift Type", "Scheme", "Departure Time", "Min Trip Duration", "Max Trip Duration"]
    write_header(wb_out_forward, schedule_header)
    write_header(wb_out_backward, schedule_header)
    # temporary destination output - BODGE
    dest_path = "output_temp.xlsx"
    wb_out.save(dest_path)

    # Bodge - inputs sheets are inconsistent

    column_locations_bus_no = get_header_coordinates(
        source_file=path,
        sheet_index=1,
        target_header= "Bus No"
    )

    print(f"Found {len(column_locations_bus_no)} columns with that name:\n")

    for index, item in enumerate(column_locations_bus_no, 1):
        print(f"Instance {index}:")
        print(f"   -> Data row: {item['data row']}")
        print("-" * 30)


    column_locations_trip_start = get_header_coordinates(
        source_file=path,
        sheet_index=1,
        target_header= "Trip Start Time"
    )

    print(f"Found {len(column_locations_trip_start)} columns with that name:\n")

    for index, item in enumerate(column_locations_trip_start, 1):
        print(f"Instance {index}:")
        print(f"   -> Data row: {item['data row']}")
        print("-" * 30)

    # Handling the extra unneeded hidden table in DR-014 and DR-014A
    if identified_route in ("DR-014", "DR-014A"):
        print("Cringe detected")

        if len(column_locations_bus_no) >= 2 and len(column_locations_trip_start) >= 2:
            column_locations_bus_no.pop(1)
            column_locations_trip_start.pop(1)
            print(f"Bus no: {column_locations_bus_no}, \n Trip no: {column_locations_trip_start}")
        else:
            print(f"Error: Expected >= 2 values, found  bus_no={len(column_locations_bus_no)}, trip_start={len(column_locations_trip_start)}")

    # Copying the columns to their proper places, checking if the route is normal or queer
    if identified_route in NORMAL_ROUTES:
        queerness = False
        bus_0_dest, bus_1_dest = 0, 1  # Normal orientation
    elif identified_route in QUEER_ROUTES:
        print("how queer...")
        queerness = True
        bus_0_dest, bus_1_dest = 1, 0  # Queer orientation
    else:
        raise ValueError("Error: Route layout rules are not defined.")

        # Execute copy function
    copy_column_as_text_between_workbooks(
        path, 1, column_locations_bus_no[0]["data row"], dest_path, bus_0_dest, "A2", False
    )
    copy_column_as_text_between_workbooks(
        path, 1, column_locations_trip_start[0]["data row"], dest_path, bus_0_dest, "D2", True
    )
    copy_column_as_text_between_workbooks(
        path, 1, column_locations_bus_no[1]["data row"], dest_path, bus_1_dest, "A2", False
    )
    copy_column_as_text_between_workbooks(
        path, 1, column_locations_trip_start[1]["data row"], dest_path, bus_1_dest, "D2", True
    )


    # Filling the rest of the rows (Morning/Evening shift, Forward/Backward, Trip duration)
    travel_time = get_travel_time(
        path,
        1,
        "Travel Time"
    )

    if identified_route in ("DR-014", "DR-014A"):
        print("Cringe detected, in time")

        if len(travel_time) >= 2:
            travel_time.pop(1)
            print(f"The travel times are {travel_time}")
        else:
            print(f"Error: Expected >= 2 values, found travel_time={len(travel_time)}")


    print(f"Here is the travel time coordinates {travel_time[0]}, {travel_time[1]}")

    if queerness == False:
        fwd_travel_time = convert_to_total_minutes(ws_details[travel_time[0]["coordinate"]].value)
        bwd_travel_time = convert_to_total_minutes(ws_details[travel_time[1]["coordinate"]].value)
        print(f"Here is the travel time data {fwd_travel_time}, {bwd_travel_time}")
    else:
        bwd_travel_time = convert_to_total_minutes(ws_details[travel_time[0]["coordinate"]].value)
        fwd_travel_time = convert_to_total_minutes(ws_details[travel_time[1]["coordinate"]].value)
        print(f"Here is the travel time data {bwd_travel_time}, {fwd_travel_time}")

    scheme_and_trip_duration(
        path,
        dest_path,
        fwd_travel_time,
        bwd_travel_time
    )

    if identified_route in ("DR-014", "DR-014A"):
        print("Cringe detected, yet again")
        autofill_next_row(dest_path)

    # Needed to make it ready for upload to system
    trim_empty_rows_and_columns(dest_path)

    # Make output folder if it doesnt exists
    os.makedirs("./Finished schedules", exist_ok=True)

    # Converting the xlsx file (cant use) to xls file and deleting the old file
    convert_xlsx_to_xls(dest_path,f"./Finished schedules/{just_the_name} - DONE.xls")
    os.remove(dest_path)



if __name__ == "__main__":
    run_excel_stuff()